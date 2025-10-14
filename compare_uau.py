from smb.SMBConnection import SMBConnection
import os
from datetime import date
from dateutil.relativedelta import relativedelta
import pandas as pd
import pyodbc
from openpyxl import load_workbook

USERNAME = os.environ.get("USERNAME_SIEG")
PASSWORD = os.environ.get("PASSWORD_SIEG")
USERNAME_SERVER = os.environ.get("USERNAME_SERVER")
PASSWORD_SERVER = os.environ.get("PASSWORD_SERVER")


today =  date.today()
year = today.strftime("%y")
FILE_NAME = ''
#checa a data de para carregar as notas do sieg
if today.day >= 20:
    DATAINIEMISSAO = date(2025,today.month,1)
    DATAFIMEMISSAO = date(2025,today.month,21)
    FILE_NAME =f"Relatorio Xml Cofre SIEG - 20.{today.month}.{year}"
else:
    DATAINIEMISSAO = date(2025,(today.month-1),1)
    DATAFIMEMISSAO = date(2025,(today.month),1)
    FILE_NAME =f"Relatorio Xml Cofre SIEG - 02.{today.month}.{year}"

SERVER_IP = "10.10.1.8"
SERVER_NAME =  'tuiuiu'
SHARE_NAME = "01 - Área Livre"
REMOTE_PATH = r"FISCAL\NFs Pendentes UAU"
TEMP_LOCAL_PATH = "/home/fernandatrentino/Ginco/temp"

# Connection string (adjust DRIVER and other parameters as needed)
conn_str = (
    f"DRIVER={'ODBC Driver 17 for SQL Server'};"
    f"SERVER={'10.203.194.150,1438'};"
    f"DATABASE={'UAUginco'};"
    f"UID={os.environ.get('USERNAME_DBUAU')};"
    f"PWD={os.environ.get('PASSWORD_DBUAU')};"
    "LoginTimeout=30;"
)

def get_data_from_sieg_xlsx(nf_type):
    
    conn = SMBConnection(
        USERNAME_SERVER, 
        PASSWORD_SERVER,
        "ubuntu_client", 
        SERVER_NAME
    )

    try:
        connected = conn.connect(SERVER_IP)
        print('Servidor remoto conectado com sucesso')
    except Exception as e:
        print(f"Erro ao conectar ao servidor {SHARE_NAME}: {e}")
        connected = False

    if connected:
        try:
            temp_file_path = f"{TEMP_LOCAL_PATH}/{FILE_NAME} - {nf_type}.xlsx"
            remote_file_path = f"{REMOTE_PATH}\{FILE_NAME} - {nf_type}.xlsx"
            with open(temp_file_path, 'wb') as f:
                conn.retrieveFile(SHARE_NAME, remote_file_path, f)

            print(f"Arquivos recuperados de {SHARE_NAME} com sucesso!")

            df = pd.read_excel(temp_file_path,sheet_name="Relátorio de Xml's - Cofre",skiprows=2)

            conn.close()
            return df
                    
        except Exception as e:
            print(f"Arquivos não foram encontrados em {SHARE_NAME} devido a {e}")  
            conn.close()

def save_file(new_file_temp_path, remote_file_path):

    conn = SMBConnection(
        USERNAME_SERVER, 
        PASSWORD_SERVER,
        "ubuntu_client", 
        SERVER_NAME
    )

    try:
        connected = conn.connect(SERVER_IP)
        print('Servidor remoto conectado com sucesso')
    except Exception as e:
        print(f"Erro ao conectar ao servidor {SHARE_NAME}: {e}")
        connected = False

    if connected:
        try:
            with open(new_file_temp_path, 'rb') as fp:
                print('Remote path:',remote_file_path)
                conn.storeFile(SHARE_NAME, remote_file_path, fp)
            
            print(f"Arquivos foram salvos em {SHARE_NAME} com sucesso!")
            conn.close()
            
            return True
        
        except Exception as e:
            print(f"Arquivos não foram salvos em {SHARE_NAME} devido a {e}")  
            conn.close()

            return False

def get_data_from_uau(list_of_filters,nf_type):
    try:
        conn = pyodbc.connect(conn_str)
        print("Conexão com o banco bem sucedida")   

        dtIni = DATAINIEMISSAO.strftime("%m-%d-%Y")
        dtFim = DATAFIMEMISSAO.strftime("%m-%d-%Y")
        print(f"""Nfs emitidas no uau com datas entre {dtIni} e {dtFim}, (%m-%d-%Y)""") 
        separator = "','"
        filter_string = separator.join(list_of_filters) 
        if nf_type == 'nfe':
            tipo_nfe = '0' #Estadual
            especie_nfe= 'NF'
        elif nf_type == 'nfse':
            tipo_nfe = '1' #Municipal
            especie_nfe = 'NF'
        else:
            tipo_nfe =0
            especie_nfe = 'CT'            


        query = f"""
            WITH notaFiscais as (
                SELECT DISTINCT
                    CONCAT(
                        p.cpf_pes,
                        e.CGC_emp,
                        SUBSTRING(nfe.NumNfAux_nfe  , PATINDEX('%[^0]%', nfe.NumNfAux_nfe +'.'), LEN(nfe.NumNfAux_nfe )),
                        REPLACE(FORMAT(nfe.ValorTotNota_nfe,'0.00'),'.','')
                    ) as chaveFilter
                    , nfe.ChaveNFe_nfe as ChaveNfe
                    , nfe.Empresa_nfe as Empresa 
                    , nfe.Obra_nfe as Obra
                    , np.Num_Proc as NumProcPgmto
                    , nfe.NumNfAux_nfe as NumNf
                    , p.cpf_pes as CNPJemiss
                    , e.CGC_emp as CNPJdest
                    , nfe.DataEmis_nfe as DataEmiss
                    , nfe.ValorTotNota_nfe as ValorNota
                    , nfe.Especie_nfe  as Especie
                    , nfe.Tipo_nfe as TipoNfe
                FROM NotasFiscaisEnt nfe
                LEFT JOIN Notfisc_Proc np ON np.NumNfe_Proc = nfe.Num_nfe and np.Empresa_proc = nfe.Empresa_nfe and np.Obra_Proc = nfe.Obra_nfe 
                INNER JOIN Empresas e ON e.Codigo_emp = nfe.EmpresaEscr_nfe 
                INNER JOIN Pessoas p  ON nfe.CodPes_nfe = p.cod_pes 
            )
            SELECT * FROM notaFiscais 
            WHERE 1=1 
            AND Especie  ='{especie_nfe}' 
            AND TipoNfe  = {tipo_nfe}
            AND DataEmiss between '{dtIni}' and '{dtFim}'
            AND chaveFilter in ('{filter_string}')
        """

        print(query)
        df = pd.read_sql(query, conn)
        print(df)
        conn.close()
        
        return df

    except pyodbc.Error as ex:
        print(f"Error connecting to SQL Server: {ex}")

    finally:
        if 'cnxn' in locals() and conn:
            conn.close()

def data_normalization(data):
    data = (
            data.str.replace('/','')
            .str.replace('-','')
            .str.replace('.','')
            .str.replace(',','')
            .str.replace('R$','')
        )
    

    return data

def compare_df():
    nf_types = ['nfe','nfse','cte'] 
    
    for type in nf_types:
        temp_file_path = f"{TEMP_LOCAL_PATH}/{FILE_NAME} - {type}.xlsx"
        remote_file_path = f"{REMOTE_PATH}\{FILE_NAME} - {type}.xlsx"

        df_sieg = get_data_from_sieg_xlsx(type)

        key_columns = ['CNPJ Emit','CNPJ Dest','Valor','Num NFe'] 
        cnpje = data_normalization(df_sieg[key_columns[0]].astype(str))
        cnpjd = data_normalization(df_sieg[key_columns[1]].astype(str))
        valor = data_normalization(df_sieg[key_columns[2]].astype(str))
        nf = data_normalization(df_sieg[key_columns[3]].astype(str))

        df_sieg["chaveFilter"] = cnpje+cnpjd+nf+valor
        df_sieg["original_index"] = df_sieg.index

        list_of_filters = df_sieg["chaveFilter"].values.tolist()
        print('Tamando da lista de chaves retornada:',len(list_of_filters))
        
        df_uau = get_data_from_uau(list_of_filters,type)

        df_filtered = df_sieg.merge(df_uau['chaveFilter'], on='chaveFilter', how='left', indicator=True)
        df_filtered = df_filtered[df_filtered['_merge'] == 'both'].drop(columns=['_merge'])

        print(df_filtered)

        wb = load_workbook(temp_file_path)
        ws = wb.active

        rows_to_delete = [4 + val for val in df_filtered["original_index"].values.tolist()]
        print('Linhas a serem deletadas:',rows_to_delete)

        deleted_nfs = wb.copy_worksheet(ws)
        deleted_nfs.title = "Nfs lançadas UAU"
        deleted_nfs.delete_rows(1, deleted_nfs.max_row)
        for row in rows_to_delete:
            dados_linha = [cell.value for cell in ws[row]] # Pega a primeira linha
            deleted_nfs.append(dados_linha)
            ws.delete_rows(row)

        
        new_file_temp_path = f"{TEMP_LOCAL_PATH}/{FILE_NAME} - {type} - A REGISTRAR.xlsx"
        print(new_file_temp_path)
        wb.save(new_file_temp_path)

        saved_file = save_file(new_file_temp_path, remote_file_path)
        if saved_file:
            for file in os.listdir(TEMP_LOCAL_PATH):
                os.remove(os.path.join(TEMP_LOCAL_PATH,file))


if __name__ == '__main__':
    compare_df()